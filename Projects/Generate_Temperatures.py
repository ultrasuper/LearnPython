# -*- coding:utf-8 -*-
from xlrd import open_workbook
from xlwt import Workbook 
import os, random,time
import openpyxl
import datetime

# my_path = r"E:\---工作资料---\体温记录-Temperature Record------\Template"
# output_workbook = Workbook(encoding="utf-8")
# output_sheet = output_workbook.add_sheet("Sheet1")
# output_sheet.write(0, 0, "Morning")
# output_sheet.write(0, 1, "Afternoon")
# tl = [36.5+0.1*x for x in range(8)]
# print(tl)
# random.seed(time.time())
# print(random.choice(tl))
# print(random.choice(tl))
# for i in range(1, 10):
# 	morning = random.choice(tl)
# 	afternoon = random.choice(tl)
# 	output_sheet.write(i, 0, morning)
# 	output_sheet.write(i, 1, afternoon)
# file_name = os.path.join(my_path, "Today's Temperature.xlsx")
# if os.path.exists(file_name):
# 	os.remove(file_name)
# output_workbook.save(file_name)


my_path = r"E:\---工作资料---\体温记录-Temperature Record------\Template"
in_name = r"Pakistan-Employees health statistics during coronavirus period.xlsx"

today = datetime.date.today().day
out_name = ''.join(["April", ' ', str(today), '-', in_name])
print(out_name)

in_path = os.path.join(my_path, in_name)
out_path = os.path.join(my_path, out_name)

tl = [36.5+0.1*x for x in range(7)]
# print(tl)
random.seed(time.time())

book = openpyxl.load_workbook(in_path)
sheet = book.get_sheet_by_name("Sheet1")
for i in range(2,11):
	random.seed(random.random() + today)
	morning = random.choice(tl)
	afternoon = random.choice(tl)
	sheet.cell(i, 4, morning)
	sheet.cell(i, 5, afternoon)
book.save(out_path)




